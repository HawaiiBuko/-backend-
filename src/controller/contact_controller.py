# 联系人控制器 - 处理联系人相关业务逻辑
from config.database import create_connection
from model.contact import Contact, ContactMethod
from sqlite3 import Error
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import io

def get_all_contacts():
    """获取所有联系人"""
    contacts = []
    conn = create_connection()
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM contacts")
            rows = cursor.fetchall()
            
            # 转换为Contact对象列表
            for row in rows:
                contact = Contact(
                    id=row[0],
                    name=row[1],
                    phone=row[2],
                    email=row[3],
                    address=row[4],
                    is_favorite=row[5]
                )
                contacts.append(contact.to_dict())
        except Error as e:
            print(f"查询联系人错误: {e}")
        finally:
            conn.close()  # 确保连接关闭
    
    return contacts

def get_favorite_contacts():
    """获取所有收藏的联系人"""
    contacts = []
    conn = create_connection()
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM contacts WHERE is_favorite = 1")
            rows = cursor.fetchall()
            
            # 转换为Contact对象列表
            for row in rows:
                contact = Contact(
                    id=row[0],
                    name=row[1],
                    phone=row[2],
                    email=row[3],
                    address=row[4],
                    is_favorite=row[5]
                )
                contacts.append(contact.to_dict())
        except Error as e:
            print(f"查询收藏联系人错误: {e}")
        finally:
            conn.close()  # 确保连接关闭
    
    return contacts

def get_contact_by_id(contact_id):
    """通过ID获取单个联系人"""
    conn = create_connection()
    contact = None
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            # 使用参数化查询防止SQL注入
            cursor.execute("SELECT * FROM contacts WHERE id = ?", (contact_id,))
            row = cursor.fetchone()  # 获取单条记录
            
            if row:
                contact = Contact(
                    id=row[0],
                    name=row[1],
                    phone=row[2],
                    email=row[3],
                    address=row[4],
                    is_favorite=row[5]
                ).to_dict()
        except Error as e:
            print(f"查询单个联系人错误: {e}")
        finally:
            conn.close()  # 确保连接关闭
    
    return contact

def get_contact_methods(contact_id):
    """获取联系人的所有联系方式"""
    methods = []
    conn = create_connection()
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM contact_methods WHERE contact_id = ?", (contact_id,))
            rows = cursor.fetchall()
            
            # 转换为ContactMethod对象列表
            for row in rows:
                method = ContactMethod(
                    id=row[0],
                    contact_id=row[1],
                    method_type=row[2],
                    method_value=row[3],
                    is_primary=row[4]
                )
                methods.append(method.to_dict())
        except Error as e:
            print(f"查询联系方式错误: {e}")
        finally:
            conn.close()
    
    return methods

def create_contact(contact_data):
    """创建新联系人"""
    conn = create_connection()
    new_contact = None
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            # 插入新联系人（返回自增ID）
            cursor.execute(
                """INSERT INTO contacts (name, phone, email, address, is_favorite) 
                   VALUES (?, ?, ?, ?, ?)""",
                (contact_data['name'], contact_data['phone'], 
                 contact_data.get('email', ''), contact_data.get('address', ''),
                 contact_data.get('is_favorite', 0))
            )
            conn.commit()  # 提交事务
            
            # 获取刚创建的联系人ID
            contact_id = cursor.lastrowid
            
            # 如果有联系方式数据，创建联系方式
            if 'methods' in contact_data and isinstance(contact_data['methods'], list):
                for method_data in contact_data['methods']:
                    create_contact_method(contact_id, method_data)
            
            new_contact = get_contact_by_id(contact_id)  # 返回完整的新联系人信息
        except Error as e:
            print(f"创建联系人错误: {e}")
            conn.rollback()  # 出错时回滚
        finally:
            conn.close()  # 确保连接关闭
    
    return new_contact

def create_contact_method(contact_id, method_data):
    """创建新的联系方式"""
    conn = create_connection()
    new_method = None
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            # 插入新联系方式
            cursor.execute(
                """INSERT INTO contact_methods (contact_id, method_type, method_value, is_primary) 
                   VALUES (?, ?, ?, ?)""",
                (contact_id, method_data['method_type'], method_data['method_value'],
                 method_data.get('is_primary', 0))
            )
            conn.commit()
            
            # 获取刚创建的联系方式ID
            method_id = cursor.lastrowid
            
            # 如果设置为主要联系方式，更新其他联系方式为非主要
            if method_data.get('is_primary', 0) == 1:
                cursor.execute(
                    "UPDATE contact_methods SET is_primary = 0 WHERE contact_id = ? AND id != ?",
                    (contact_id, method_id)
                )
                conn.commit()
                
            # 获取完整的联系方式信息
            cursor.execute("SELECT * FROM contact_methods WHERE id = ?", (method_id,))
            row = cursor.fetchone()
            if row:
                new_method = ContactMethod(
                    id=row[0],
                    contact_id=row[1],
                    method_type=row[2],
                    method_value=row[3],
                    is_primary=row[4]
                ).to_dict()
        except Error as e:
            print(f"创建联系方式错误: {e}")
            conn.rollback()
        finally:
            conn.close()
    
    return new_method

def update_contact(contact_id, contact_data):
    """更新已有联系人"""
    conn = create_connection()
    updated_contact = None
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            # 更新联系人信息
            cursor.execute(
                """UPDATE contacts SET name = ?, phone = ?, email = ?, address = ?, is_favorite = ? 
                   WHERE id = ?""",
                (contact_data['name'], contact_data['phone'], 
                 contact_data.get('email', ''), contact_data.get('address', ''),
                 contact_data.get('is_favorite', 0), 
                 contact_id)
            )
            conn.commit()  # 提交事务
            
            # 验证更新结果（查询更新后的联系人）
            updated_contact = get_contact_by_id(contact_id)
        except Error as e:
            print(f"更新联系人错误: {e}")
            conn.rollback()  # 出错时回滚
        finally:
            conn.close()  # 确保连接关闭
    
    return updated_contact

def update_contact_method(method_id, method_data):
    """更新联系方式"""
    conn = create_connection()
    updated_method = None
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            
            # 获取联系方式的contact_id
            cursor.execute("SELECT contact_id FROM contact_methods WHERE id = ?", (method_id,))
            result = cursor.fetchone()
            if not result:
                return None
            contact_id = result[0]
            
            # 更新联系方式
            cursor.execute(
                """UPDATE contact_methods SET method_type = ?, method_value = ?, is_primary = ? 
                   WHERE id = ?""",
                (method_data['method_type'], method_data['method_value'],
                 method_data.get('is_primary', 0), method_id)
            )
            
            # 如果设置为主要联系方式，更新其他联系方式为非主要
            if method_data.get('is_primary', 0) == 1:
                cursor.execute(
                    "UPDATE contact_methods SET is_primary = 0 WHERE contact_id = ? AND id != ?",
                    (contact_id, method_id)
                )
            
            conn.commit()
            
            # 获取更新后的联系方式
            cursor.execute("SELECT * FROM contact_methods WHERE id = ?", (method_id,))
            row = cursor.fetchone()
            if row:
                updated_method = ContactMethod(
                    id=row[0],
                    contact_id=row[1],
                    method_type=row[2],
                    method_value=row[3],
                    is_primary=row[4]
                ).to_dict()
        except Error as e:
            print(f"更新联系方式错误: {e}")
            conn.rollback()
        finally:
            conn.close()
    
    return updated_method

def delete_contact_method(method_id):
    """删除联系方式"""
    conn = create_connection()
    success = False
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM contact_methods WHERE id = ?", (method_id,))
            conn.commit()
            success = cursor.rowcount > 0
        except Error as e:
            print(f"删除联系方式错误: {e}")
            conn.rollback()
        finally:
            conn.close()
    
    return success

def toggle_favorite(contact_id):
    """切换联系人的收藏状态"""
    conn = create_connection()
    updated_contact = None
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            
            # 获取当前收藏状态
            cursor.execute("SELECT is_favorite FROM contacts WHERE id = ?", (contact_id,))
            result = cursor.fetchone()
            if not result:
                return None
            
            current_favorite = result[0]
            new_favorite = 1 if current_favorite == 0 else 0
            
            # 更新收藏状态
            cursor.execute("UPDATE contacts SET is_favorite = ? WHERE id = ?", (new_favorite, contact_id))
            conn.commit()
            
            updated_contact = get_contact_by_id(contact_id)
        except Error as e:
            print(f"切换收藏状态错误: {e}")
            conn.rollback()
        finally:
            conn.close()
    
    return updated_contact

def export_contacts_to_excel():
    """将所有联系人导出为Excel文件"""
    # 创建Workbook对象
    wb = Workbook()
    ws = wb.active
    ws.title = "联系人列表"
    
    # 设置表头
    headers = ["ID", "姓名", "主要电话", "电子邮箱", "地址", "是否收藏", "其他联系方式"]
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}1"] = header
    
    # 获取所有联系人
    contacts = get_all_contacts()
    
    # 填充数据
    for row_idx, contact in enumerate(contacts, 2):
        # 基本信息
        ws[f"A{row_idx}"] = contact["id"]
        ws[f"B{row_idx}"] = contact["name"]
        ws[f"C{row_idx}"] = contact["phone"]
        ws[f"D{row_idx}"] = contact["email"]
        ws[f"E{row_idx}"] = contact["address"]
        ws[f"F{row_idx}"] = "是" if contact["is_favorite"] == 1 else "否"
        
        # 其他联系方式
        methods = get_contact_methods(contact["id"])
        other_methods = []
        for method in methods:
            if method["method_type"] != "phone" or method["method_value"] != contact["phone"]:
                other_methods.append(f"{method['method_type']}: {method['method_value']}")
        
        ws[f"G{row_idx}"] = "; ".join(other_methods) if other_methods else "无"
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # 将Excel文件保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def import_contacts_from_excel(file_stream):
    """从Excel文件导入联系人数据"""
    try:
        # 加载Excel文件
        workbook = load_workbook(file_stream)
        worksheet = workbook.active
        
        # 读取表头
        headers = [cell.value for cell in worksheet[1]]
        
        # 从第二行开始读取数据
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # 跳过空行（检查ID列或姓名列是否为空）
            if not row[0] and not row[1]:  # ID和姓名都为空，跳过该行
                continue
            
            # 根据导出的列顺序读取数据：
            # A列(0): ID - 导入时跳过，使用数据库自增ID
            # B列(1): 姓名
            # C列(2): 主要电话
            # D列(3): 电子邮箱
            # E列(4): 地址
            # F列(5): 是否收藏
            # G列(6): 其他联系方式
            
            # 创建联系人基本信息
            contact_data = {
                'name': row[1] if len(row) > 1 and row[1] else '',
                'phone': row[2] if len(row) > 2 and row[2] else '',
                'email': row[3] if len(row) > 3 and row[3] else '',
                'address': row[4] if len(row) > 4 and row[4] else '',
                'is_favorite': 1 if len(row) > 5 and row[5] == '是' else 0
            }
            
            # 验证必填字段
            if not contact_data['name'] or not contact_data['phone']:
                continue
            
            # 创建联系人
            new_contact = create_contact(contact_data)
            if not new_contact:
                continue
            contact_id = new_contact['id']
            
            # 读取其他联系方式（G列，索引6）
            if len(row) > 6 and row[6] and row[6] != '无':
                # 解析其他联系方式，格式为 "type: value; type: value"
                other_methods_str = str(row[6])
                if other_methods_str.strip():
                    # 按分号分割多个联系方式
                    method_pairs = other_methods_str.split(';')
                    for method_pair in method_pairs:
                        method_pair = method_pair.strip()
                        if ':' in method_pair:
                            parts = method_pair.split(':', 1)
                            method_type = parts[0].strip()
                            method_value = parts[1].strip() if len(parts) > 1 else ''
                            
                            if method_type and method_value:
                                method_data = {
                                    'method_type': method_type,
                                    'method_value': method_value,
                                    'is_primary': 0  # 其他联系方式不是主要联系方式
                                }
                                # 创建联系方式
                                create_contact_method(contact_id, method_data)
        
        return True
    except Exception as e:
        print(f"导入失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def delete_contact(contact_id):
    """删除联系人"""
    conn = create_connection()
    success = False  # 标记是否删除成功
    
    if conn is not None:
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM contacts WHERE id = ?", (contact_id,))
            conn.commit()  # 提交事务
            
            # 检查是否有记录被删除（rowcount > 0表示删除成功）
            success = cursor.rowcount > 0
        except Error as e:
            print(f"删除联系人错误: {e}")
            conn.rollback()  # 出错时回滚
        finally:
            conn.close()  # 确保连接关闭
    
    return success